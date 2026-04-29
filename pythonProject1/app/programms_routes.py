from flask import Blueprint, request, jsonify, send_file
from app.database import get_db_connection
from app.utils import token_required, log_action, admin_token_required
from app.validators import Validators
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

programms_bp = Blueprint('programms', __name__)


@programms_bp.route('/programms', methods=['GET'])
@token_required
def get_programms(current_user):
    """Получение списка программ со школами и сотрудниками"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    offset = (page - 1) * per_page

    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

    with get_db_connection() as conn:
        cur = conn.cursor()

        if user_role == 'admin':
            cur.execute('SELECT COUNT(*) FROM programms')
            total = cur.fetchone()['count']

            cur.execute('''
                SELECT 
                    p.id,
                    p.name,
                    p.comment,
                    p.year,
                    p.user_id,
                    p.created_at,
                    u.login as owner_login,
                    u.surname as owner_surname,
                    u.name as owner_name
                FROM programms p
                JOIN users u ON p.user_id = u.id
                ORDER BY p.year DESC, p.created_at DESC
                LIMIT %s OFFSET %s
            ''', (per_page, offset))
        else:
            cur.execute('SELECT COUNT(*) FROM programms WHERE user_id = %s', (user_id,))
            total = cur.fetchone()['count']

            cur.execute('''
                SELECT 
                    p.id,
                    p.name,
                    p.comment,
                    p.year,
                    p.user_id,
                    p.created_at,
                    u.login as owner_login,
                    u.surname as owner_surname,
                    u.name as owner_name
                FROM programms p
                JOIN users u ON p.user_id = u.id
                WHERE p.user_id = %s
                ORDER BY p.year DESC, p.created_at DESC
                LIMIT %s OFFSET %s
            ''', (user_id, per_page, offset))

        programms = cur.fetchall()

        # Для каждой программы получаем школы и сотрудников
        for program in programms:
            if program.get('created_at'):
                program['created_at'] = program['created_at'].strftime('%Y-%m-%d %H:%M:%S')

            # Получаем школы для программы
            cur.execute('''
                SELECT DISTINCT 
                    s.id, s.name, s.comment
                FROM program_school_employees pse
                JOIN schools s ON pse.school_id = s.id
                WHERE pse.program_id = %s
                ORDER BY s.name
            ''', (program['id'],))
            schools = cur.fetchall()

            # Для каждой школы получаем сотрудников
            for school in schools:
                cur.execute('''
                    SELECT DISTINCT u.id, u.surname, u.name, u.patronymic, u.email
                    FROM program_school_employees pse
                    JOIN users u ON pse.employee_id = u.id
                    WHERE pse.program_id = %s 
                        AND pse.school_id = %s 
                        AND u.role = 'employeer'
                        AND pse.employee_id IS NOT NULL
                    ORDER BY u.surname, u.name
                ''', (program['id'], school['id']))
                school['employees'] = cur.fetchall() or []

            program['schools'] = schools or []

        return jsonify({
            'programms': programms,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page,
            'is_admin': user_role == 'admin'
        })


@programms_bp.route('/programms/export', methods=['GET'])
@admin_token_required
def export_programms(current_user):
    """Экспорт всех программ в XLSX файл с информацией о школах и сотрудниках"""
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()

            # Получаем все программы с их школами и сотрудниками
            cur.execute('''
                SELECT 
                    p.id,
                    p.name as program_name,
                    p.comment as program_comment,
                    p.year,
                    u.login as owner_login,
                    u.surname as owner_surname,
                    u.name as owner_name,
                    p.created_at
                FROM programms p
                JOIN users u ON p.user_id = u.id
                ORDER BY p.year DESC, p.created_at DESC
            ''')
            programs = cur.fetchall()

            # Собираем данные для Excel
            data = []
            for program in programs:
                # Получаем школы для программы
                cur.execute('''
                    SELECT DISTINCT 
                        s.id,
                        s.name as school_name,
                        s.comment as school_comment,
                        STRING_AGG(DISTINCT CONCAT(u.surname, ' ', u.name, COALESCE(' ' || u.patronymic, '')), '; ') as employees
                    FROM program_school_employees pse
                    JOIN schools s ON pse.school_id = s.id
                    LEFT JOIN users u ON pse.employee_id = u.id AND u.role = 'employeer'
                    WHERE pse.program_id = %s
                    GROUP BY s.id, s.name, s.comment
                    ORDER BY s.name
                ''', (program['id'],))
                schools = cur.fetchall()

                if not schools:
                    # Если нет школ, добавляем одну строку без школ
                    data.append({
                        'Название программы': program['program_name'],
                        'Комментарий к программе': program['program_comment'] or '',
                        'Год': program['year'] if program['year'] else '',
                        'Школа': '',
                        'Комментарий к школе': '',
                        'Сотрудники': ''
                    })
                else:
                    # Для каждой школы добавляем отдельную строку
                    for school in schools:
                        data.append({
                            'Название программы': program['program_name'],
                            'Комментарий к программе': program['program_comment'] or '',
                            'Год': program['year'] if program['year'] else '',
                            'Школа': school['school_name'] or '',
                            'Комментарий к школе': school['school_comment'] or '',
                            'Сотрудники': school['employees'] or ''
                        })

        # Создаем DataFrame
        df = pd.DataFrame(data)

        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Программы', index=False)

            # Настраиваем стили
            worksheet = writer.sheets['Программы']

            # Жирный шрифт для заголовков
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C4A27A", end_color="C4A27A", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Автоматическая ширина колонок
            worksheet.column_dimensions['A'].width = 30  # Название программы
            worksheet.column_dimensions['B'].width = 40  # Комментарий к программе
            worksheet.column_dimensions['C'].width = 10  # Год
            worksheet.column_dimensions['D'].width = 30  # Школа
            worksheet.column_dimensions['E'].width = 40  # Комментарий к школе
            worksheet.column_dimensions['F'].width = 40  # Сотрудники

            # Добавляем фильтр
            worksheet.auto_filter.ref = worksheet.dimensions

        output.seek(0)

        log_action(current_user['userId'], 'EXPORT_PROGRAMMS',
                   f"Exported {len(programs)} programms with schools to Excel")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='programms.xlsx'
        )

    except Exception as e:
        log_action(current_user['userId'], 'EXPORT_PROGRAMMS_ERROR', str(e))
        return jsonify({'error': f'Ошибка экспорта: {str(e)}'}), 500


@programms_bp.route('/programms/template', methods=['GET'])
@admin_token_required
def download_template(current_user):
    """Скачивание шаблона для импорта программ"""
    try:
        # Создаем шаблон только с заголовками (без примеров)
        df = pd.DataFrame({
            'Название программы': [],
            'Комментарий': [],
            'Год': []
        })

        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Шаблон программ', index=False)

            # Настраиваем стили
            worksheet = writer.sheets['Шаблон программ']

            # Жирный шрифт для заголовков
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C4A27A", end_color="C4A27A", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Автоматическая ширина колонок
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 15

        output.seek(0)

        log_action(current_user['userId'], 'DOWNLOAD_PROGRAMMS_TEMPLATE',
                   "Downloaded programms import template")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='programms_template.xlsx'
        )

    except Exception as e:
        log_action(current_user['userId'], 'DOWNLOAD_TEMPLATE_ERROR', str(e))
        return jsonify({'error': f'Ошибка скачивания шаблона: {str(e)}'}), 500


@programms_bp.route('/programms/import', methods=['POST'])
@admin_token_required
def import_programms(current_user):
    """Импорт программ из XLSX файла"""
    try:
        # Проверяем, есть ли файл в запросе
        if 'file' not in request.files:
            return jsonify({'error': 'Файл не выбран'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'Файл не выбран'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Неверный формат файла. Поддерживаются только .xlsx и .xls'}), 400

        # Читаем Excel файл
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': f'Ошибка чтения файла: {str(e)}'}), 400

        # Проверяем, что файл не пустой
        if df.empty:
            return jsonify({'error': 'Файл пустой'}), 400

        # Проверяем наличие необходимых колонок
        required_columns = ['Название программы']
        for col in required_columns:
            if col not in df.columns:
                return jsonify({'error': f'В файле отсутствует колонка "{col}"'}), 400

        # Удаляем пустые строки по названию программы
        df = df.dropna(subset=['Название программы'])

        if df.empty:
            return jsonify({'error': 'Файл не содержит данных для импорта'}), 400

        # Проверяем на дубликаты названий программ
        program_names = df['Название программы'].astype(str).str.strip()
        duplicates = program_names[program_names.duplicated()].unique()

        if len(duplicates) > 0:
            return jsonify({
                'error': f'В файле найдены дубликаты названий программ: {", ".join(duplicates[:5])}'
                         + ('...' if len(duplicates) > 5 else '')
            }), 400

        # Проверяем существующие программы в БД
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute('SELECT name FROM programms')
            existing_programms = [row['name'] for row in cur.fetchall()]

        # Находим дубликаты с существующими программами
        existing_duplicates = [name for name in program_names if name in existing_programms]

        if len(existing_duplicates) > 0:
            return jsonify({
                'error': f'Следующие программы уже существуют в базе: {", ".join(existing_duplicates[:5])}'
                         + ('...' if len(existing_duplicates) > 5 else '')
            }), 400

        # Импортируем программы
        imported_count = 0
        errors = []

        with get_db_connection() as conn:
            cur = conn.cursor()

            for idx, row in df.iterrows():
                try:
                    name = str(row['Название программы']).strip()
                    comment = str(row['Комментарий']).strip() if pd.notna(row.get('Комментарий')) else None
                    year = row.get('Год')

                    # Валидация года
                    if pd.notna(year):
                        try:
                            year_int = int(year)
                            if year_int < 1900 or year_int > 2100:
                                errors.append(f'Строка {idx + 2}: Год "{year}" должен быть в диапазоне 1900-2100')
                                continue
                            year = year_int
                        except ValueError:
                            errors.append(f'Строка {idx + 2}: Год "{year}" должен быть числом')
                            continue
                    else:
                        year = None

                    # Валидация названия
                    if len(name) < 2:
                        errors.append(f'Строка {idx + 2}: Название "{name}" слишком короткое (мин. 2 символа)')
                        continue

                    if len(name) > 200:
                        errors.append(f'Строка {idx + 2}: Название "{name}" слишком длинное (макс. 200 символов)')
                        continue

                    if comment and len(comment) > 1000:
                        errors.append(f'Строка {idx + 2}: Комментарий слишком длинный (макс. 1000 символов)')
                        continue

                    # Вставляем программу
                    cur.execute('''
                        INSERT INTO programms (name, comment, year, user_id)
                        VALUES (%s, %s, %s, %s)
                    ''', (name, comment, year, current_user['userId']))

                    imported_count += 1

                except Exception as e:
                    errors.append(f'Строка {idx + 2}: {str(e)}')

        log_action(current_user['userId'], 'IMPORT_PROGRAMMS',
                   f"Imported {imported_count} programms, errors: {len(errors)}")

        if errors:
            return jsonify({
                'success': True,
                'message': f'Импортировано {imported_count} программ. Ошибки: {"; ".join(errors[:5])}',
                'imported_count': imported_count,
                'errors': errors
            }), 207  # Multi-Status

        return jsonify({
            'success': True,
            'message': f'Успешно импортировано {imported_count} программ',
            'imported_count': imported_count
        }), 200

    except Exception as e:
        log_action(current_user['userId'], 'IMPORT_PROGRAMMS_ERROR', str(e))
        return jsonify({'error': f'Ошибка импорта: {str(e)}'}), 500


@programms_bp.route('/programms', methods=['POST'])
@token_required
def create_program(current_user):
    """Создание новой программы с привязкой школ через program_school_employees"""
    data = request.json
    name = Validators.sanitize_input(data.get('name'))
    comment = Validators.sanitize_input(data.get('comment'))
    year = data.get('year')
    school_ids = data.get('school_ids', [])

    if not name:
        return jsonify({'error': 'Название программы обязательно'}), 400

    if len(name) < 2:
        return jsonify({'error': 'Название должно содержать минимум 2 символа'}), 400

    if len(name) > 200:
        return jsonify({'error': 'Название не может быть длиннее 200 символов'}), 400

    if comment and len(comment) > 1000:
        return jsonify({'error': 'Комментарий не может быть длиннее 1000 символов'}), 400

    if year:
        try:
            year_int = int(year)
            if year_int < 1900 or year_int > 2100:
                return jsonify({'error': 'Год должен быть в диапазоне 1900-2100'}), 400
        except ValueError:
            return jsonify({'error': 'Год должен быть числом'}), 400

    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute('''
                INSERT INTO programms (name, comment, year, user_id)
                VALUES (%s, %s, %s, %s)
                RETURNING id, name, comment, year, user_id, created_at
            ''', (name, comment, year, current_user['userId']))

            new_program = cur.fetchone()
            program_id = new_program['id']

            # Привязываем школы - создаем записи в program_school_employees с NULL employee_id
            for school_id in school_ids:
                cur.execute('''
                    INSERT INTO program_school_employees (program_id, school_id, employee_id)
                    VALUES (%s, %s, NULL)
                ''', (program_id, school_id))

            if new_program and new_program.get('created_at'):
                new_program['created_at'] = new_program['created_at'].strftime('%Y-%m-%d %H:%M:%S')

            new_program['schools'] = []

            log_action(current_user['userId'], 'CREATE_PROGRAM',
                       f"Created program: {name} (Year: {year}) with {len(school_ids)} schools")

            return jsonify(new_program), 201

    except Exception as e:
        log_action(current_user['userId'], 'CREATE_PROGRAM_ERROR', str(e))
        return jsonify({'error': 'Внутренняя ошибка сервера'}), 500


@programms_bp.route('/programms/<int:program_id>', methods=['GET'])
@token_required
def get_program(current_user, program_id):
    """Получение одной программы со школами и сотрудниками"""
    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()

        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        cur.execute('''
            SELECT 
                p.id,
                p.name,
                p.comment,
                p.year,
                p.user_id,
                p.created_at,
                u.login as owner_login,
                u.surname as owner_surname,
                u.name as owner_name
            FROM programms p
            JOIN users u ON p.user_id = u.id
            WHERE p.id = %s
        ''', (program_id,))
        program = cur.fetchone()

        if not program:
            return jsonify({'error': 'Программа не найдена'}), 404

        if user_role != 'admin' and program['user_id'] != user_id:
            return jsonify({'error': 'Доступ запрещен'}), 403

        if program.get('created_at'):
            program['created_at'] = program['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        # Получаем школы для программы
        cur.execute('''
            SELECT DISTINCT 
                s.id, s.name, s.comment
            FROM program_school_employees pse
            JOIN schools s ON pse.school_id = s.id
            WHERE pse.program_id = %s
            ORDER BY s.name
        ''', (program_id,))
        schools = cur.fetchall()

        # Для каждой школы получаем сотрудников
        for school in schools:
            cur.execute('''
                SELECT DISTINCT u.id, u.surname, u.name, u.patronymic, u.email
                FROM program_school_employees pse
                JOIN users u ON pse.employee_id = u.id
                WHERE pse.program_id = %s 
                    AND pse.school_id = %s 
                    AND u.role = 'employeer'
                    AND pse.employee_id IS NOT NULL
                ORDER BY u.surname, u.name
            ''', (program_id, school['id']))
            school['employees'] = cur.fetchall() or []

        program['schools'] = schools or []

        return jsonify(program)


@programms_bp.route('/programms/<int:program_id>', methods=['PUT'])
@token_required
def update_program(current_user, program_id):
    """Обновление программы и её школ через program_school_employees"""
    data = request.json
    name = Validators.sanitize_input(data.get('name'))
    comment = Validators.sanitize_input(data.get('comment'))
    year = data.get('year')
    school_ids = data.get('school_ids', [])

    if not name:
        return jsonify({'error': 'Название программы обязательно'}), 400

    if len(name) < 2:
        return jsonify({'error': 'Название должно содержать минимум 2 символа'}), 400

    if len(name) > 200:
        return jsonify({'error': 'Название не может быть длиннее 200 символов'}), 400

    if comment and len(comment) > 1000:
        return jsonify({'error': 'Комментарий не может быть длиннее 1000 символов'}), 400

    if year:
        try:
            year_int = int(year)
            if year_int < 1900 or year_int > 2100:
                return jsonify({'error': 'Год должен быть в диапазоне 1900-2100'}), 400
        except ValueError:
            return jsonify({'error': 'Год должен быть числом'}), 400

    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()

        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        cur.execute('SELECT user_id FROM programms WHERE id = %s', (program_id,))
        program = cur.fetchone()

        if not program:
            return jsonify({'error': 'Программа не найдена'}), 404

        if user_role != 'admin' and program['user_id'] != user_id:
            return jsonify({'error': 'Доступ запрещен'}), 403

        # Обновляем программу
        cur.execute('''
            UPDATE programms
            SET name = %s, comment = %s, year = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, name, comment, year, user_id, created_at
        ''', (name, comment, year if year else None, program_id))

        updated_program = cur.fetchone()

        # Обновляем школы программы через program_school_employees
        if school_ids is not None:
            # Получаем текущие привязанные школы
            cur.execute('''
                SELECT DISTINCT school_id FROM program_school_employees WHERE program_id = %s
            ''', (program_id,))
            current_schools = [row['school_id'] for row in cur.fetchall()]

            # Определяем какие школы добавить и какие удалить
            schools_to_add = set(school_ids) - set(current_schools)
            schools_to_remove = set(current_schools) - set(school_ids)

            # Удаляем школы
            for school_id in schools_to_remove:
                cur.execute('''
                    DELETE FROM program_school_employees 
                    WHERE program_id = %s AND school_id = %s
                ''', (program_id, school_id))

            # Добавляем новые школы
            for school_id in schools_to_add:
                cur.execute('''
                    INSERT INTO program_school_employees (program_id, school_id, employee_id)
                    VALUES (%s, %s, NULL)
                ''', (program_id, school_id))

        if updated_program and updated_program.get('created_at'):
            updated_program['created_at'] = updated_program['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        log_action(current_user['userId'], 'UPDATE_PROGRAM',
                   f"Updated program ID {program_id}: {name} (Year: {year})")

        return jsonify(updated_program)


@programms_bp.route('/programms/<int:program_id>', methods=['DELETE'])
@token_required
def delete_program(current_user, program_id):
    """Удаление программы"""
    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()

        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        cur.execute('SELECT user_id, name FROM programms WHERE id = %s', (program_id,))
        program = cur.fetchone()

        if not program:
            return jsonify({'error': 'Программа не найдена'}), 404

        if user_role != 'admin' and program['user_id'] != user_id:
            return jsonify({'error': 'Доступ запрещен'}), 403

        # Удаляем связи с школами (автоматически каскадно удалятся из-за ON DELETE CASCADE)
        cur.execute('DELETE FROM program_school_employees WHERE program_id = %s', (program_id,))

        # Удаляем программу
        cur.execute('DELETE FROM programms WHERE id = %s', (program_id,))

        log_action(current_user['userId'], 'DELETE_PROGRAM',
                   f"Deleted program ID {program_id}: {program['name']}")

        return jsonify({'success': True, 'message': 'Программа удалена'})