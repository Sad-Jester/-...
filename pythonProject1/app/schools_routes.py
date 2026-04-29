from flask import Blueprint, request, jsonify, send_file
from app.database import get_db_connection
from app.utils import token_required, log_action, admin_token_required
from app.validators import Validators
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment

schools_bp = Blueprint('schools', __name__)


@schools_bp.route('/schools', methods=['GET'])
@token_required
def get_schools(current_user):
    """Получение списка школ с привязанными программами и сотрудниками"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    offset = (page - 1) * per_page

    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

    with get_db_connection() as conn:
        cur = conn.cursor()

        if user_role == 'admin':
            cur.execute('SELECT COUNT(*) FROM schools')
            total = cur.fetchone()['count']

            cur.execute('''
                SELECT 
                    s.id, s.name, s.comment, s.user_id, s.created_at,
                    u.login as owner_login, u.surname as owner_surname, u.name as owner_name
                FROM schools s
                JOIN users u ON s.user_id = u.id
                ORDER BY s.created_at DESC
                LIMIT %s OFFSET %s
            ''', (per_page, offset))
        else:
            cur.execute('SELECT COUNT(*) FROM schools WHERE user_id = %s', (user_id,))
            total = cur.fetchone()['count']

            cur.execute('''
                SELECT 
                    s.id, s.name, s.comment, s.user_id, s.created_at,
                    u.login as owner_login, u.surname as owner_surname, u.name as owner_name
                FROM schools s
                JOIN users u ON s.user_id = u.id
                WHERE s.user_id = %s
                ORDER BY s.created_at DESC
                LIMIT %s OFFSET %s
            ''', (user_id, per_page, offset))

        schools = cur.fetchall()

        # Для каждой школы получаем привязанные программы
        for school in schools:
            if school.get('created_at'):
                school['created_at'] = school['created_at'].strftime('%Y-%m-%d %H:%M:%S')

            # Получаем программы для школы
            cur.execute('''
                SELECT DISTINCT 
                    p.id, p.name, p.year, p.comment
                FROM program_school_employees pse
                JOIN programms p ON pse.program_id = p.id
                WHERE pse.school_id = %s
                ORDER BY p.year DESC, p.name
            ''', (school['id'],))
            programs = cur.fetchall()

            # Для каждой программы получаем сотрудников
            for program in programs:
                cur.execute('''
                    SELECT DISTINCT u.id, u.surname, u.name, u.patronymic, u.email
                    FROM program_school_employees pse
                    JOIN users u ON pse.employee_id = u.id
                    WHERE pse.program_id = %s 
                        AND pse.school_id = %s 
                        AND u.role = 'employeer'
                        AND pse.employee_id IS NOT NULL
                    ORDER BY u.surname, u.name
                ''', (program['id'], school['id']))
                program['employees'] = cur.fetchall() or []

            school['programs'] = programs or []

        return jsonify({
            'schools': schools,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page,
            'is_admin': user_role == 'admin'
        })


@schools_bp.route('/schools', methods=['POST'])
@token_required
def create_school(current_user):
    """Создание новой школы с привязкой к программам"""
    data = request.json

    name = Validators.sanitize_input(data.get('name'))
    comment = Validators.sanitize_input(data.get('comment'))
    program_ids = data.get('program_ids', [])

    if not name:
        return jsonify({'error': 'Название школы обязательно'}), 400

    if len(name) < 2:
        return jsonify({'error': 'Название должно содержать минимум 2 символа'}), 400

    if len(name) > 200:
        return jsonify({'error': 'Название не может быть длиннее 200 символов'}), 400

    if comment and len(comment) > 1000:
        return jsonify({'error': 'Комментарий не может быть длиннее 1000 символов'}), 400

    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute('''
                INSERT INTO schools (name, comment, user_id)
                VALUES (%s, %s, %s)
                RETURNING id, name, comment, user_id, created_at
            ''', (name, comment, current_user['userId']))

            new_school = cur.fetchone()
            school_id = new_school['id']

            # Привязываем программы - создаем записи в program_school_employees с NULL employee_id
            for program_id in program_ids:
                cur.execute('''
                    INSERT INTO program_school_employees (program_id, school_id, employee_id)
                    VALUES (%s, %s, NULL)
                ''', (program_id, school_id))

            if new_school and new_school.get('created_at'):
                new_school['created_at'] = new_school['created_at'].strftime('%Y-%m-%d %H:%M:%S')

            new_school['programs'] = []

            log_action(current_user['userId'], 'CREATE_SCHOOL',
                       f"Created school: {name} with {len(program_ids)} programs")

            return jsonify(new_school), 201

    except Exception as e:
        log_action(current_user['userId'], 'CREATE_SCHOOL_ERROR', str(e))
        return jsonify({'error': 'Внутренняя ошибка сервера'}), 500


@schools_bp.route('/schools/<int:school_id>', methods=['GET'])
@token_required
def get_school(current_user, school_id):
    """Получение одной школы с привязанными программами и сотрудниками"""
    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()

        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        cur.execute('''
            SELECT 
                s.id, s.name, s.comment, s.user_id, s.created_at,
                u.login as owner_login, u.surname as owner_surname, u.name as owner_name
            FROM schools s
            JOIN users u ON s.user_id = u.id
            WHERE s.id = %s
        ''', (school_id,))
        school = cur.fetchone()

        if not school:
            return jsonify({'error': 'Школа не найдена'}), 404

        if user_role != 'admin' and school['user_id'] != user_id:
            return jsonify({'error': 'Доступ запрещен'}), 403

        if school.get('created_at'):
            school['created_at'] = school['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        # Получаем программы для школы
        cur.execute('''
            SELECT DISTINCT 
                p.id, p.name, p.year, p.comment
            FROM program_school_employees pse
            JOIN programms p ON pse.program_id = p.id
            WHERE pse.school_id = %s
            ORDER BY p.year DESC, p.name
        ''', (school_id,))
        programs = cur.fetchall()

        # Для каждой программы получаем сотрудников
        for program in programs:
            cur.execute('''
                SELECT DISTINCT u.id, u.surname, u.name, u.patronymic, u.email
                FROM program_school_employees pse
                JOIN users u ON pse.employee_id = u.id
                WHERE pse.program_id = %s 
                    AND pse.school_id = %s 
                    AND u.role = 'employeer'
                    AND pse.employee_id IS NOT NULL
                ORDER BY u.surname, u.name
            ''', (program['id'], school_id))
            program['employees'] = cur.fetchall() or []

        school['programs'] = programs or []

        return jsonify(school)


@schools_bp.route('/schools/<int:school_id>', methods=['PUT'])
@token_required
def update_school(current_user, school_id):
    """Обновление школы и привязанных программ"""
    data = request.json

    name = Validators.sanitize_input(data.get('name'))
    comment = Validators.sanitize_input(data.get('comment'))
    program_ids = data.get('program_ids', [])

    if not name:
        return jsonify({'error': 'Название школы обязательно'}), 400

    if len(name) < 2:
        return jsonify({'error': 'Название должно содержать минимум 2 символа'}), 400

    if len(name) > 200:
        return jsonify({'error': 'Название не может быть длиннее 200 символов'}), 400

    if comment and len(comment) > 1000:
        return jsonify({'error': 'Комментарий не может быть длиннее 1000 символов'}), 400

    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()

        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        cur.execute('SELECT user_id FROM schools WHERE id = %s', (school_id,))
        school = cur.fetchone()

        if not school:
            return jsonify({'error': 'Школа не найдена'}), 404

        if user_role != 'admin' and school['user_id'] != user_id:
            return jsonify({'error': 'Доступ запрещен'}), 403

        cur.execute('''
            UPDATE schools
            SET name = %s, comment = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, name, comment, user_id, created_at
        ''', (name, comment, school_id))

        updated_school = cur.fetchone()

        # Получаем текущие привязанные программы (уникальные program_id)
        cur.execute('''
            SELECT DISTINCT program_id FROM program_school_employees WHERE school_id = %s
        ''', (school_id,))
        current_programs = [row['program_id'] for row in cur.fetchall()]

        # Определяем какие программы добавить и какие удалить
        programs_to_add = set(program_ids) - set(current_programs)
        programs_to_remove = set(current_programs) - set(program_ids)

        # Удаляем программы, которые больше не привязаны (удаляем все записи с этой школой и программой)
        for program_id in programs_to_remove:
            cur.execute('''
                DELETE FROM program_school_employees 
                WHERE program_id = %s AND school_id = %s
            ''', (program_id, school_id))

        # Добавляем новые программы
        for program_id in programs_to_add:
            cur.execute('''
                INSERT INTO program_school_employees (program_id, school_id, employee_id)
                VALUES (%s, %s, NULL)
            ''', (program_id, school_id))

        if updated_school and updated_school.get('created_at'):
            updated_school['created_at'] = updated_school['created_at'].strftime('%Y-%m-%d %H:%M:%S')

        log_action(current_user['userId'], 'UPDATE_SCHOOL', f"Updated school ID {school_id}: {name}")

        return jsonify(updated_school)


@schools_bp.route('/schools/<int:school_id>', methods=['DELETE'])
@token_required
def delete_school(current_user, school_id):
    """Удаление школы"""
    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()

        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        cur.execute('SELECT user_id, name FROM schools WHERE id = %s', (school_id,))
        school = cur.fetchone()

        if not school:
            return jsonify({'error': 'Школа не найдена'}), 404

        if user_role != 'admin' and school['user_id'] != user_id:
            return jsonify({'error': 'Доступ запрещен'}), 403

        # Удаляем связи с программами
        cur.execute('DELETE FROM program_school_employees WHERE school_id = %s', (school_id,))

        # Удаляем школу
        cur.execute('DELETE FROM schools WHERE id = %s', (school_id,))

        log_action(current_user['userId'], 'DELETE_SCHOOL', f"Deleted school ID {school_id}: {school['name']}")

        return jsonify({'success': True, 'message': 'Школа удалена'})


@schools_bp.route('/schools/available-programs', methods=['GET'])
@token_required
def get_available_programs(current_user):
    """Получение списка всех программ для привязки к школе"""
    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        if user_role == 'admin':
            cur.execute('''
                SELECT id, name, year, comment
                FROM programms
                ORDER BY year DESC, name
            ''')
        else:
            cur.execute('''
                SELECT id, name, year, comment
                FROM programms
                WHERE user_id = %s
                ORDER BY year DESC, name
            ''', (user_id,))

        programs = cur.fetchall()
        return jsonify(programs)


# Экспорт школ
@schools_bp.route('/schools/export', methods=['GET'])
@admin_token_required
def export_schools(current_user):
    """Экспорт всех школ в XLSX файл с информацией о программах и сотрудниках"""
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()

            # Получаем все школы
            cur.execute('''
                SELECT 
                    s.id,
                    s.name as school_name,
                    s.comment as school_comment,
                    u.login as owner_login,
                    u.surname as owner_surname,
                    u.name as owner_name,
                    s.created_at
                FROM schools s
                JOIN users u ON s.user_id = u.id
                ORDER BY s.created_at DESC
            ''')
            schools = cur.fetchall()

            # Собираем данные для Excel
            data = []
            for school in schools:
                # Получаем программы для школы с сотрудниками
                cur.execute('''
                    SELECT DISTINCT 
                        p.id,
                        p.name as program_name,
                        p.comment as program_comment,
                        p.year,
                        STRING_AGG(DISTINCT CONCAT(u.surname, ' ', u.name, COALESCE(' ' || u.patronymic, '')), '; ') as employees
                    FROM program_school_employees pse
                    JOIN programms p ON pse.program_id = p.id
                    LEFT JOIN users u ON pse.employee_id = u.id AND u.role = 'employeer'
                    WHERE pse.school_id = %s
                    GROUP BY p.id, p.name, p.comment, p.year
                    ORDER BY p.year DESC, p.name
                ''', (school['id'],))
                programs = cur.fetchall()

                if not programs:
                    # Если нет программ, добавляем одну строку без программ
                    data.append({
                        'Название школы': school['school_name'],
                        'Комментарий к школе': school['school_comment'] or '',
                        'Программа': '',
                        'Комментарий к программе': '',
                        'Год программы': '',
                        'Сотрудники': ''
                    })
                else:
                    # Для каждой программы добавляем отдельную строку
                    for program in programs:
                        data.append({
                            'Название школы': school['school_name'],
                            'Комментарий к школе': school['school_comment'] or '',
                            'Программа': program['program_name'] or '',
                            'Комментарий к программе': program['program_comment'] or '',
                            'Год программы': program['year'] if program['year'] else '',
                            'Сотрудники': program['employees'] or ''
                        })

        # Создаем DataFrame
        df = pd.DataFrame(data)

        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Школы', index=False)

            # Настраиваем стили
            worksheet = writer.sheets['Школы']

            # Жирный шрифт для заголовков
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C4A27A", end_color="C4A27A", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Автоматическая ширина колонок
            worksheet.column_dimensions['A'].width = 30  # Название школы
            worksheet.column_dimensions['B'].width = 40  # Комментарий к школе
            worksheet.column_dimensions['C'].width = 30  # Программа
            worksheet.column_dimensions['D'].width = 40  # Комментарий к программе
            worksheet.column_dimensions['E'].width = 12  # Год программы
            worksheet.column_dimensions['F'].width = 40  # Сотрудники

            # Добавляем фильтр
            worksheet.auto_filter.ref = worksheet.dimensions

        output.seek(0)

        log_action(current_user['userId'], 'EXPORT_SCHOOLS',
                   f"Exported {len(schools)} schools with programs to Excel")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='schools.xlsx'
        )

    except Exception as e:
        log_action(current_user['userId'], 'EXPORT_SCHOOLS_ERROR', str(e))
        return jsonify({'error': f'Ошибка экспорта: {str(e)}'}), 500


# Скачивание шаблона
@schools_bp.route('/schools/template', methods=['GET'])
@admin_token_required
def download_template(current_user):
    """Скачивание шаблона для импорта школ"""
    try:
        df = pd.DataFrame({'Название школы': [], 'Комментарий': []})
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Шаблон школ', index=False)
            worksheet = writer.sheets['Шаблон школ']
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C4A27A", end_color="C4A27A", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 50

        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='schools_template.xlsx')
    except Exception as e:
        return jsonify({'error': f'Ошибка: {str(e)}'}), 500


# Импорт школ
@schools_bp.route('/schools/import', methods=['POST'])
@admin_token_required
def import_schools(current_user):
    """Импорт школ из XLSX файла"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Файл не выбран'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Файл не выбран'}), 400
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Неверный формат файла'}), 400

        df = pd.read_excel(file)
        if df.empty:
            return jsonify({'error': 'Файл пустой'}), 400

        if 'Название школы' not in df.columns:
            return jsonify({'error': 'Отсутствует колонка "Название школы"'}), 400

        df = df.dropna(subset=['Название школы'])
        if df.empty:
            return jsonify({'error': 'Нет данных для импорта'}), 400

        imported_count = 0
        errors = []

        with get_db_connection() as conn:
            cur = conn.cursor()
            for idx, row in df.iterrows():
                try:
                    name = str(row['Название школы']).strip()
                    comment = str(row['Комментарий']).strip() if pd.notna(row.get('Комментарий')) else None

                    if len(name) < 2:
                        errors.append(f'Строка {idx + 2}: название слишком короткое')
                        continue
                    if len(name) > 200:
                        errors.append(f'Строка {idx + 2}: название слишком длинное')
                        continue

                    cur.execute('INSERT INTO schools (name, comment, user_id) VALUES (%s, %s, %s)',
                                (name, comment, current_user['userId']))
                    imported_count += 1
                except Exception as e:
                    errors.append(f'Строка {idx + 2}: {str(e)}')

        log_action(current_user['userId'], 'IMPORT_SCHOOLS', f"Imported {imported_count} schools")
        return jsonify(
            {'success': True, 'message': f'Импортировано {imported_count} школ', 'imported_count': imported_count})

    except Exception as e:
        return jsonify({'error': f'Ошибка импорта: {str(e)}'}), 500

@schools_bp.route('/employeers/list', methods=['GET'])
@token_required
def get_employeers_list(current_user):
    """Получение списка пользователей с ролью employeer"""
    user_id = current_user['userId']
    user_role = None

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()
        user_role = user_row['role'] if user_row else 'user'

        # Только админ может получать список всех employeer
        if user_role != 'admin':
            return jsonify({'error': 'Доступ запрещен'}), 403

        cur.execute('''
            SELECT id, surname, name, patronymic, login, email
            FROM users
            WHERE role = 'employeer'
            ORDER BY surname, name
        ''')
        users = cur.fetchall()

        # Форматируем ФИО
        for user in users:
            full_name = f"{user['surname']} {user['name']}"
            if user.get('patronymic'):
                full_name += f" {user['patronymic']}"
            user['full_name'] = full_name

        return jsonify(users)